"""ResultExporter — Excel, Markdown, and LaTeX export for SuiteResult."""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

from .aggregator import AggregatedLeaf, AggregatedMetrics
from .config import RunnerConfig
from .errors import ExportError

if TYPE_CHECKING:
    from pathlib import Path

    from .suite import SuiteResult


def _flatten_aggregated(metrics: AggregatedMetrics, prefix: str = "") -> dict[str, AggregatedLeaf]:
    """Flatten nested AggregatedMetrics to {dotted_path: AggregatedLeaf}."""
    out: dict[str, AggregatedLeaf] = {}
    for key, value in metrics.items():
        full_key = f"{prefix}.{key}" if prefix else key
        if isinstance(value, AggregatedLeaf):
            out[full_key] = value
        elif isinstance(value, dict):
            out.update(_flatten_aggregated(value, full_key))
    return out


def _flatten_results(
    suite_result: SuiteResult,
) -> tuple[
    dict[str, dict[str, AggregatedLeaf]],  # scalar_metrics[path][exp_name]
    dict[str, dict[str, AggregatedLeaf]],  # sequence_metrics[path][exp_name]
]:
    """Separate scalar and sequence metrics across all experiments."""
    scalar: dict[str, dict[str, AggregatedLeaf]] = {}
    sequence: dict[str, dict[str, AggregatedLeaf]] = {}

    for exp_name, agg_metrics in suite_result.results.items():
        flat = _flatten_aggregated(agg_metrics)
        for path, leaf in flat.items():
            if isinstance(leaf.mean, list):
                sequence.setdefault(path, {})[exp_name] = leaf
            else:
                scalar.setdefault(path, {})[exp_name] = leaf

    return scalar, sequence


def _is_better(a: float, b: float, direction: str) -> bool:
    if direction == "higher_is_better":
        return a > b
    return a < b


def _best_exp(
    path: str,
    exp_leaves: dict[str, AggregatedLeaf],
    cfg: RunnerConfig,
) -> str | None:
    direction = cfg.metric_directions.get(path, cfg.default_metric_direction)
    best_name: str | None = None
    best_val: float | None = None
    for name, leaf in exp_leaves.items():
        if not isinstance(leaf.mean, float):
            continue
        if best_val is None or _is_better(leaf.mean, best_val, direction):
            best_val = leaf.mean
            best_name = name
    return best_name


class ResultExporter:
    def __init__(self, runner_config: RunnerConfig | None = None) -> None:
        self.cfg = runner_config or RunnerConfig()

    def export(self, suite_result: SuiteResult, output_dir: Path) -> None:
        try:
            for fmt in self.cfg.export_formats:
                if fmt == "markdown":
                    self._export_markdown(suite_result, output_dir)
                elif fmt == "excel":
                    self._export_excel(suite_result, output_dir)
                elif fmt == "latex":
                    self._export_latex(suite_result, output_dir)
        except Exception as exc:
            msg = f"Export failed: {exc}"
            raise ExportError(msg) from exc

    # ------------------------------------------------------------------
    # Markdown
    # ------------------------------------------------------------------

    def _export_markdown(self, suite_result: SuiteResult, output_dir: Path) -> None:
        from latent_space.utils.markdown_results import render_markdown_table

        scalar_metrics, sequence_metrics = _flatten_results(suite_result)
        exp_names = suite_result.experiments
        lines: list[str] = [f"# {self.cfg.suite_name}\n"]

        # Scalar table
        if scalar_metrics:
            lines.append("## Scalar Metrics\n")
            headers = ["metric", *exp_names]
            rows = []
            for path, exp_leaves in scalar_metrics.items():
                row: dict[str, Any] = {"metric": path}
                for exp in exp_names:
                    leaf = exp_leaves.get(exp)
                    row[exp] = f"{leaf.mean:.4f} ± {leaf.std:.4f}" if leaf is not None else ""
                rows.append(row)
            lines.append(render_markdown_table(headers, rows))
            lines.append("")

        # Sequence sections
        if sequence_metrics:
            lines.append("## Sequence Metrics\n")
            for path, exp_leaves in sequence_metrics.items():
                lines.append(f"### {path}\n")
                mode = self.cfg.sequence_export_mode

                if mode == "full" or self.cfg.markdown_full_sequences:
                    # Full table: one row per step
                    max_len = max(
                        len(leaf.mean)  # type: ignore[arg-type]
                        for leaf in exp_leaves.values()
                        if isinstance(leaf.mean, list)
                    )
                    headers = (
                        ["step"]
                        + [f"{e}.mean" for e in exp_names]
                        + [f"{e}.std" for e in exp_names]
                    )
                    rows = []
                    for i in range(max_len):
                        row = {"step": i}
                        for exp in exp_names:
                            leaf = exp_leaves.get(exp)
                            if leaf and isinstance(leaf.mean, list) and i < len(leaf.mean):
                                row[f"{exp}.mean"] = f"{leaf.mean[i]:.4f}"
                                row[f"{exp}.std"] = f"{leaf.std[i]:.4f}"  # type: ignore[index]
                            else:
                                row[f"{exp}.mean"] = ""
                                row[f"{exp}.std"] = ""
                        rows.append(row)
                    lines.append(render_markdown_table(headers, rows))
                elif mode == "final":
                    headers = ["experiment", "final_mean", "final_std"]
                    rows = []
                    for exp in exp_names:
                        leaf = exp_leaves.get(exp)
                        if leaf and isinstance(leaf.mean, list):
                            rows.append(
                                {
                                    "experiment": exp,
                                    "final_mean": f"{leaf.mean[-1]:.4f}",
                                    "final_std": f"{leaf.std[-1]:.4f}",  # type: ignore[index]
                                },
                            )
                    lines.append(render_markdown_table(headers, rows))
                else:  # summary
                    headers = ["experiment", "first", "last", "best"]
                    rows = []
                    for exp in exp_names:
                        leaf = exp_leaves.get(exp)
                        if leaf and isinstance(leaf.mean, list) and leaf.mean:
                            direction = self.cfg.metric_directions.get(
                                path,
                                self.cfg.default_metric_direction,
                            )
                            best = (
                                max(leaf.mean)
                                if direction == "higher_is_better"
                                else min(leaf.mean)
                            )
                            rows.append(
                                {
                                    "experiment": exp,
                                    "first": f"{leaf.mean[0]:.4f}",
                                    "last": f"{leaf.mean[-1]:.4f}",
                                    "best": f"{best:.4f}",
                                },
                            )
                    lines.append(render_markdown_table(headers, rows))
                lines.append("")

        (output_dir / f"{self.cfg.suite_name}.md").write_text("\n".join(lines), encoding="utf-8")

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------

    def _export_excel(self, suite_result: SuiteResult, output_dir: Path) -> None:
        import openpyxl
        from openpyxl.styles import PatternFill

        scalar_metrics, sequence_metrics = _flatten_results(suite_result)
        exp_names = suite_result.experiments
        wb = openpyxl.Workbook()

        # Scalar sheet
        ws = wb.active
        ws.title = "Scalar Metrics"
        highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        header_row = ["metric", *exp_names]
        ws.append(header_row)

        for path, exp_leaves in scalar_metrics.items():
            best_name = _best_exp(path, exp_leaves, self.cfg)
            row_data = [path]
            for exp in exp_names:
                leaf = exp_leaves.get(exp)
                row_data.append(f"{leaf.mean:.4f} ± {leaf.std:.4f}" if leaf is not None else "")
            ws.append(row_data)
            # Highlight best cell
            if best_name and best_name in exp_names:
                col_idx = exp_names.index(best_name) + 2  # 1-based + metric col
                ws.cell(row=ws.max_row, column=col_idx).fill = highlight

        # Sequence sheets (one per metric)
        for path, exp_leaves in sequence_metrics.items():
            sheet_name = path[:31]
            seq_ws = wb.create_sheet(title=sheet_name)
            col_headers = ["step"]
            for exp in exp_names:
                col_headers += [f"{exp}.mean", f"{exp}.std"]
            seq_ws.append(col_headers)

            max_len = max(
                len(leaf.mean)  # type: ignore[arg-type]
                for leaf in exp_leaves.values()
                if isinstance(leaf.mean, list)
            )
            for i in range(max_len):
                row = [i]
                for exp in exp_names:
                    leaf = exp_leaves.get(exp)
                    if leaf and isinstance(leaf.mean, list) and i < len(leaf.mean):
                        row += [leaf.mean[i], leaf.std[i]]  # type: ignore[index]
                    else:
                        row += ["", ""]
                seq_ws.append(row)

        wb.save(output_dir / f"{self.cfg.suite_name}.xlsx")

    # ------------------------------------------------------------------
    # LaTeX
    # ------------------------------------------------------------------

    def _export_latex(self, suite_result: SuiteResult, output_dir: Path) -> None:
        scalar_metrics, sequence_metrics = _flatten_results(suite_result)
        exp_names = suite_result.experiments
        dp = self.cfg.latex_decimal_places
        fmt = f"{{:.{dp}f}}"

        blocks: list[str] = []

        # Scalar table
        if scalar_metrics:
            col_spec = "l" + "c" * len(exp_names)
            header = " & ".join(["Metric", *exp_names]) + r" \\"
            rows_lines: list[str] = []
            for path, exp_leaves in scalar_metrics.items():
                best_name = _best_exp(path, exp_leaves, self.cfg)
                cells = [path.replace("_", r"\_")]
                for exp in exp_names:
                    leaf = exp_leaves.get(exp)
                    if leaf is None:
                        cells.append("")
                    else:
                        cell = f"{fmt.format(leaf.mean)} \\pm {fmt.format(leaf.std)}"
                        if exp == best_name:
                            cell = r"\textbf{" + cell + "}"
                        cells.append(f"${cell}$")
                rows_lines.append(" & ".join(cells) + r" \\")

            table = (
                "\\begin{table}[htbp]\n"
                "\\centering\n"
                "\\begin{tabular}{" + col_spec + "}\n"
                "\\toprule\n" + header + "\n"
                "\\midrule\n" + "\n".join(rows_lines) + "\n"
                "\\bottomrule\n"
                "\\end{tabular}\n"
                f"\\caption{{{self.cfg.suite_name} scalar metrics}}\n"
                "\\end{table}"
            )
            blocks.append(table)

        # Sequence tables
        for path, exp_leaves in sequence_metrics.items():
            col_spec = "c" + "cc" * len(exp_names)
            sub_headers = []
            for exp in exp_names:
                sub_headers += [f"{exp} mean", f"{exp} std"]
            header = " & ".join(["Step", *sub_headers]) + r" \\"

            max_len = max(
                len(leaf.mean)  # type: ignore[arg-type]
                for leaf in exp_leaves.values()
                if isinstance(leaf.mean, list)
            )
            rows_lines = []
            for i in range(max_len):
                cells = [str(i)]
                for exp in exp_names:
                    leaf = exp_leaves.get(exp)
                    if leaf and isinstance(leaf.mean, list) and i < len(leaf.mean):
                        cells += [
                            fmt.format(leaf.mean[i]),
                            fmt.format(leaf.std[i]),  # type: ignore[index]
                        ]
                    else:
                        cells += ["", ""]
                rows_lines.append(" & ".join(cells) + r" \\")

            safe_path = path.replace("_", r"\_")
            table = (
                "\\begin{table}[htbp]\n"
                "\\centering\n"
                "\\begin{tabular}{" + col_spec + "}\n"
                "\\toprule\n" + header + "\n"
                "\\midrule\n" + "\n".join(rows_lines) + "\n"
                "\\bottomrule\n"
                "\\end{tabular}\n"
                f"\\caption{{Sequence metric: {safe_path}}}\n"
                "\\end{table}"
            )
            blocks.append(table)

        (output_dir / f"{self.cfg.suite_name}.tex").write_text(
            "\n\n".join(blocks) + "\n",
            encoding="utf-8",
        )
